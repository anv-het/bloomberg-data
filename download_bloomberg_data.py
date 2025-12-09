"""
Bloomberg Financial Data Downloader
====================================
This script downloads financial data from Bloomberg Terminal for Indian stocks.

Requirements:
- Bloomberg Terminal must be installed and running
- Excel must be installed
- Bloomberg Excel template file: FA1_vwijagme.xlsx

Usage:
    python download_bloomberg_data.py --symbol HDFCB
    python download_bloomberg_data.py --symbol IOCL --output_dir ./output
"""

import os
import time
import sys
import argparse
import subprocess
from pathlib import Path
from datetime import datetime
import pandas as pd

try:
    from openpyxl import load_workbook
    import xlwings as xw
except ImportError as e:
    print(f"❌ Missing required package: {e}")
    print("Please install requirements: pip install openpyxl xlwings pandas")
    sys.exit(1)


class BloombergDataDownloader:
    def __init__(self, template_path: str, output_dir: str = "./output"):
        """
        Initialize the Bloomberg Data Downloader
        
        Args:
            template_path: Path to Bloomberg Excel template (FA1_vwijagme.xlsx)
            output_dir: Directory to save output files
        """
        self.template_path = Path(template_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True, parents=True)
        
        # Default symbol in template (the one to replace)
        self.default_symbol = "IOCL"
        
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
    
    def replace_symbol_in_template(self, new_symbol: str, temp_file_path: Path) -> None:
        """
        Replace the default symbol with a new symbol in the Excel template
        
        Args:
            new_symbol: Bloomberg symbol (e.g., HDFCB, IOCL)
            temp_file_path: Path where modified template will be saved
        """
        print(f"📝 Replacing {self.default_symbol} with {new_symbol} in template...")
        
        wb = load_workbook(self.template_path)
        ws = wb["BBG Adj Highlights"]
        
        # Go through every row (skip header row)
        replacements = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str):
                    if self.default_symbol in cell.value:
                        old_value = cell.value
                        cell.value = cell.value.replace(self.default_symbol, new_symbol)
                        print(f"   ✓ {old_value} → {cell.value}")
                        replacements += 1
        
        wb.save(temp_file_path)
        print(f"✅ Made {replacements} replacements. Saved to: {temp_file_path}")
    
    def refresh_bloomberg_data(self, xlsx_path: Path, output_excel: Path, 
                               output_values: Path, wait_seconds: int = 15) -> bool:
        """
        Open Excel file to refresh Bloomberg formulas and save results
        
        Args:
            xlsx_path: Path to Excel file with Bloomberg formulas
            output_excel: Path to save Excel with formulas
            output_values: Path to save values-only Excel
            wait_seconds: Seconds to wait for Bloomberg data to refresh
        
        Returns:
            True if successful, False otherwise
        """
        print(f"\n🔄 Opening Excel to refresh Bloomberg data...")
        print(f"⏳ Will wait {wait_seconds} seconds for data refresh...")
        
        try:
            # Open the file in Excel
            os.startfile(str(xlsx_path))
            time.sleep(wait_seconds)
            
            # Connect to the active Excel application
            app = xw.apps.active
            if app is None:
                raise RuntimeError('❌ No Excel application found')
            
            # Find the workbook
            book = None
            for b in app.books:
                try:
                    if xlsx_path.name in b.name:
                        book = b
                        break
                except:
                    pass
            
            if book is None:
                raise RuntimeError('❌ Workbook not found in Excel')
            
            print("📊 Creating values-only copy...")
            
            # Create a new workbook for values
            vals_book = app.books.add()
            
            # Copy each sheet's data as values
            for sh in book.sheets:
                print(f"   Copying sheet: {sh.name}")
                
                # Create or get target sheet
                try:
                    target = vals_book.sheets[sh.name]
                except Exception:
                    target = vals_book.sheets.add(sh.name)
                
                # Copy values
                data = sh.used_range.value
                target.range("A1").value = data
            
            # Remove default empty sheet if present
            try:
                vals_book.sheets["Sheet1"].delete()
            except Exception:
                pass
            
            # Save both files
            print(f"💾 Saving values-only file: {output_values}")
            vals_book.save(str(output_values))
            vals_book.close()
            
            print(f"💾 Saving Excel file: {output_excel}")
            book.save(str(output_excel))
            book.close()
            
            print("✅ Files saved successfully!")
            return True
            
        except Exception as e:
            print(f"❌ Error: {e}")
            return False
            
        finally:
            # Clean up Excel
            if 'app' in locals() and app is not None:
                app.quit()
            
            # Force kill Excel if still running
            try:
                subprocess.run(["taskkill", "/IM", "EXCEL.EXE", "/F"], 
                             capture_output=True, timeout=5)
            except:
                pass
    
    def export_to_csv(self, excel_path: Path, csv_path: Path) -> None:
        """
        Export the main data sheet to CSV
        
        Args:
            excel_path: Path to Excel file
            csv_path: Path to save CSV file
        """
        print(f"\n📄 Exporting to CSV: {csv_path}")
        
        try:
            # Read the main data sheet
            df = pd.read_excel(excel_path, sheet_name="BBG Adj Highlights")
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"✅ CSV exported successfully!")
        except Exception as e:
            print(f"⚠️  CSV export failed: {e}")
    
    def download_data(self, symbol: str, wait_seconds: int = 15) -> dict:
        """
        Main method to download Bloomberg data for a symbol
        
        Args:
            symbol: Bloomberg symbol (e.g., HDFCB, IOCL)
            wait_seconds: Seconds to wait for Bloomberg refresh
        
        Returns:
            Dictionary with paths to output files
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # File paths
        temp_file = self.output_dir / f"temp_{symbol}_{timestamp}.xlsx"
        output_excel = self.output_dir / f"{symbol}_bloomberg_data_{timestamp}.xlsx"
        output_values = self.output_dir / f"{symbol}_bloomberg_values_{timestamp}.xlsx"
        output_csv = self.output_dir / f"{symbol}_bloomberg_data_{timestamp}.csv"
        
        print(f"\n{'='*60}")
        print(f"📊 Bloomberg Data Downloader")
        print(f"{'='*60}")
        print(f"Symbol: {symbol}")
        print(f"Output Directory: {self.output_dir}")
        print(f"{'='*60}\n")
        
        # Step 1: Replace symbol in template
        self.replace_symbol_in_template(symbol, temp_file)
        
        # Step 2: Refresh Bloomberg data
        success = self.refresh_bloomberg_data(
            temp_file, output_excel, output_values, wait_seconds
        )
        
        if not success:
            print("\n❌ Failed to download Bloomberg data")
            return None
        
        # Step 3: Export to CSV
        self.export_to_csv(output_values, output_csv)
        
        # Clean up temp file
        try:
            temp_file.unlink()
        except:
            pass
        
        print(f"\n{'='*60}")
        print(f"✅ SUCCESS! Bloomberg data downloaded for {symbol}")
        print(f"{'='*60}")
        print(f"📁 Output files:")
        print(f"   Excel (with formulas): {output_excel}")
        print(f"   Excel (values only):   {output_values}")
        print(f"   CSV:                   {output_csv}")
        print(f"{'='*60}\n")
        
        return {
            'symbol': symbol,
            'excel': output_excel,
            'values': output_values,
            'csv': output_csv
        }


def main():
    parser = argparse.ArgumentParser(
        description='Download financial data from Bloomberg Terminal',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Download data for HDFCB
  python download_bloomberg_data.py --symbol HDFCB
  
  # Download data for IOCL with custom output directory
  python download_bloomberg_data.py --symbol IOCL --output_dir ./my_data
  
  # Download with custom template and longer wait time
  python download_bloomberg_data.py --symbol RELIANCE --template C:/custom/template.xlsx --wait 30
        """
    )
    
    parser.add_argument('--symbol', '-s', required=True, 
                       help='Bloomberg symbol (e.g., HDFCB, IOCL, RELIANCE)')
    parser.add_argument('--template', '-t', 
                       default=r'C:\blp\data\FA1_vwijagme.xlsx',
                       help='Path to Bloomberg Excel template (default: C:/blp/data/FA1_vwijagme.xlsx)')
    parser.add_argument('--output_dir', '-o', 
                       default='./output',
                       help='Output directory (default: ./output)')
    parser.add_argument('--wait', '-w', type=int, default=15,
                       help='Seconds to wait for Bloomberg refresh (default: 15)')
    
    args = parser.parse_args()
    
    try:
        downloader = BloombergDataDownloader(
            template_path=args.template,
            output_dir=args.output_dir
        )
        
        result = downloader.download_data(
            symbol=args.symbol.upper(),
            wait_seconds=args.wait
        )
        
        if result:
            sys.exit(0)
        else:
            sys.exit(1)
            
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
